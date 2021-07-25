import openpyxl
import os
import time 
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import xlrd

options = Options()
options.add_argument('--headless')
options.add_argument('--hide-scrollbars')



ID = 'hoge'
PW = 'hogehoge'
URL = 'hogehoge'



driver = webdriver.Chrome(chrome_options=options)
url = 'https://{}:{}@{}'.format(ID, PW, URL)
print(url)
driver.get(url)


## Excelから値を取得
book = xlrd.open_workbook("master.xlsx")
sheet = book.sheet_by_name("Sheet1")




##空のエクセルナンバーリスト作成
NUMBERLIST = []

"""
##空のタイトルリスト作成
TITLELIST = []
"""

##空のURLリスト作成
URLLIST = []

##空のナンバーリストにナンバーをいれる
for row in range(sheet.nrows):
    NUMBER = sheet.cell(row,0).value
    NUMBERLIST.append(int(NUMBER))



## 空のURLリストにURLをいれる
for row in range(sheet.nrows):
    URL = sheet.cell(row,1).value
    URLLIST.append(URL)

    ##URLを開く
    driver.get(URL)

    
    ##PC用のサイズ取得
    page_height = driver.execute_script('return document.body.scrollHeight')
    driver.set_window_size(1200, page_height)
    
    #1秒待機 

    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(3)
    ##タイトルリストの値でファイルの命名
    FILENAME = os.path.join(str(NUMBERLIST[row]) + ".png")
    driver.save_screenshot(FILENAME)
[print(i) for i in NUMBERLIST]

driver.quit()


## SPキャプチャ

##UAの偽装
options.add_argument('--user-agent=Mozilla/5.0 (iPhone; CPU iPhone OS 14_0_1 like Mac OS X) AppleWebKit/602.3.12 (KHTML, like Gecko) Version/10.0 Mobile/14C92 Safari/602.1')

## ブラウザを起動
driver = webdriver.Chrome(chrome_options=options)
options.add_argument('--hide-scrollbars')

SPURL = 'hogehoge'

url = 'https://{}:{}@{}'.format(ID, PW, SPURL)
print(url)
driver.get(url)

## 空のURLリストにURLをいれる
for row in range(sheet.nrows):
    URL = sheet.cell(row,1).value
    URLLIST.append(URL)


    ##URLを開く
    driver.get(URL)

    ##リロード
    driver.refresh()

    #3秒待機 
    time.sleep(4)
    
    ##SP用のサイズ取得
    page_height = driver.execute_script('return document.body.scrollHeight')
    driver.set_window_size(414, page_height)    
    


    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    
    ##タイトルリストの値でファイルの命名
    FILENAME = os.path.join("【SP】"+str(NUMBERLIST[row]) + ".png")
    driver.save_screenshot(FILENAME)
[print(i) for i in NUMBERLIST]


driver.quit()

wb = openpyxl.load_workbook("コーディング指示書_サンプルv4.xlsm", keep_vba=True)
ws = wb.worksheets


for (sheet_name, url) in zip(NUMBERLIST, URLLIST):
    sheet_name = int(sheet_name)
    copied_sheet = wb.copy_worksheet(wb["コーディング指示書サンプル"])
    copied_sheet.title = str(sheet_name)
    copied_sheet.cell(row = 1, column = 2).value = sheet_name
    copied_sheet.cell(row = 1, column = 4).value = url


wb.save("コーディング指示書_サンプルv4.xlsm")

print(len(ws))

wb = openpyxl.load_workbook("コーディング指示書_サンプルv4.xlsm", keep_vba=True)
ws = wb.worksheets

ws.pop(0)##1－ト目のディレクトリマップを削除する


print(len(ws))

count = 0
for i in ws:
    IMAGE_PC = str(NUMBERLIST[count]) + '.png'
    IMAGE_SP = "【SP】"+str(NUMBERLIST[count]) + '.png'
    sheet =ws[count]
    img_file_pc = os.path.abspath(IMAGE_PC)
    img_file_sp = os.path.abspath(IMAGE_SP)
    img_pc = openpyxl.drawing.image.Image(img_file_pc)
    img_sp = openpyxl.drawing.image.Image(img_file_sp)
    sheet.add_image( img_pc, "B6" )
    sheet.add_image( img_sp, "AF6" )
    count = count + 1


wb.save("コーディング指示書_サンプルv4.xlsm")


print("指示書作成しました")

    