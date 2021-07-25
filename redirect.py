import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time 
import csv
import pandas as pd
import openpyxl
import re
import xlrd


options = Options()
options.add_argument('--headless')
options.add_argument('--hide-scrollbars')


ID = 'hoge'
PW = 'hogehoge'
URL = 'hogehoge'



driver = webdriver.Chrome(chrome_options=options)
basicurl = 'https://{}:{}@{}'.format(ID, PW, URL)
print(basicurl)
driver.get(basicurl)


book = xlrd.open_workbook("master.xlsx")
sheet = book.sheet_by_name("Sheet1")

URLS = []



wb = openpyxl.Workbook()
sheets = wb.active
sheets.title = 'PHPサーバーリダイレクト洗い出し'
count = 1
urlcount = 1

for row in range(sheet.nrows):
    URL = sheet.cell(row,1).value
    URLS.append(URL)
    print(URL)
    driver.get(URL)
    time.sleep(2)
    elements = driver.find_elements_by_tag_name("a")




    URLLIST = []

    for i in elements:
        datas = i.get_attribute("href")
        URLLIST.append(datas)
        title = driver.title
        print(datas)
        sheets.cell(column=1, row=count, value=title)
        sheets.cell(column=2, row=count, value=datas)
        count = count + 1
        print(count)


    for urls in URLLIST:
        driver.get(urls)
        cur_url = driver.current_url
        sheets.cell(column=3, row=urlcount, value=cur_url)
        urlcount = urlcount + 1
        print(urlcount)



wb.save('scraping_excel.xlsx')
wb.close()

